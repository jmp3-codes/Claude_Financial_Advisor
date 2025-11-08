#!/usr/bin/env python3
"""
Create Excel master files for Financial Advisor
Run from terminal: python3 create_master_files.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

base_path = os.path.dirname(os.path.abspath(__file__))

# Create merchant_dictionary.xlsx
wb_merchant = openpyxl.Workbook()
ws_merchant = wb_merchant.active
ws_merchant.title = "Merchants"

merchant_headers = ["Raw Statement Text", "Business Name", "Category", "Subcategory", "Notes", "Last Seen", "Frequency"]
ws_merchant.append(merchant_headers)

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws_merchant[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

ws_merchant.column_dimensions['A'].width = 25
ws_merchant.column_dimensions['B'].width = 25
ws_merchant.column_dimensions['C'].width = 15
ws_merchant.column_dimensions['D'].width = 15
ws_merchant.column_dimensions['E'].width = 30
ws_merchant.column_dimensions['F'].width = 12
ws_merchant.column_dimensions['G'].width = 10

merchant_path = os.path.join(base_path, "merchant_dictionary.xlsx")
wb_merchant.save(merchant_path)
print(f"✅ Created: {merchant_path}")

# Create master_transactions_2025.xlsx with multiple tabs
wb_trans = openpyxl.Workbook()
wb_trans.remove(wb_trans.active)

ws_all_trans = wb_trans.create_sheet("All Transactions", 0)
trans_headers = ["Transaction ID", "Date", "Account Source", "Raw Description", "Business Name", 
                 "Category", "Subcategory", "Amount", "Transaction Type", "Statement Period", 
                 "Calendar Month", "Notes", "Verified"]
ws_all_trans.append(trans_headers)

for cell in ws_all_trans[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws_all_trans.column_dimensions['A'].width = 18
ws_all_trans.column_dimensions['B'].width = 12
ws_all_trans.column_dimensions['C'].width = 15
ws_all_trans.column_dimensions['D'].width = 25
ws_all_trans.column_dimensions['E'].width = 20
ws_all_trans.column_dimensions['F'].width = 15
ws_all_trans.column_dimensions['G'].width = 15
ws_all_trans.column_dimensions['H'].width = 12
ws_all_trans.column_dimensions['I'].width = 12
ws_all_trans.column_dimensions['J'].width = 18
ws_all_trans.column_dimensions['K'].width = 14
ws_all_trans.column_dimensions['L'].width = 15
ws_all_trans.column_dimensions['M'].width = 10

ws_monthly = wb_trans.create_sheet("Monthly Summary", 1)
ws_monthly.append(["Month", "Total Income", "Total Expenses", "Net Cash Flow", "Housing", "Transportation", 
                   "Food & Dining", "Healthcare", "Entertainment", "Shopping", "Personal Care", 
                   "Financial", "Insurance", "Subscriptions", "Travel", "Education", "Gifts & Donations", 
                   "Pets", "Other"])

for cell in ws_monthly[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws_categories = wb_trans.create_sheet("Category Analysis", 2)
ws_categories.append(["Category", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "YTD Total", "Monthly Avg"])

for cell in ws_categories[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

ws_dashboard = wb_trans.create_sheet("Annual Dashboard", 3)
ws_dashboard['A1'] = "2025 Financial Summary"
ws_dashboard['A1'].font = Font(bold=True, size=14)

trans_path = os.path.join(base_path, "master_transactions_2025.xlsx")
wb_trans.save(trans_path)
print(f"✅ Created: {trans_path}")

print("\n✅ All files created successfully!")
print("You can now see them in Finder at:")
print("/Users/jackpage/Documents/Coding/Claude/Financial Advisor/master_files/")
